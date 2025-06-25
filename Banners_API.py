import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from rembg import remove
import requests
import zipfile
import io
import re

st.set_page_config(layout="wide")
st.title("Banner Formatter (KVI, Range, Dual MRP) + All_MbIDs Tab + Images")

st.markdown("""
**Required Excel Sheet Columns:**

- `Hubs | Product Name | Focused Sub Cat | MB ID 1 | MB ID 2 (Range, Dual MRP)`
And Seperate tabs for KVI, Range & Dual MRP banners

Please ensure your Excel sheet contains these columns with the exact names for proper processing.
""")

ALLOWED_HUBS = [
    "ggn", "jpr", "ahm", "ind", "mum", "pun", "blr", "hyd", "chn",
    "dlwss", "pbssnurpur", "krssdavanagere", "mgss", "west bengal",
    "vikaspuri & rp bagh", "nurpur & vr mall", "SS North", "SS South"
]

def clean_mbid(val):
    try:
        num = float(val)
        if num.is_integer():
            return str(int(num))
        return str(val)
    except Exception:
        return str(val).strip()
        
def is_hub(val):
    if not isinstance(val, str):
        return False
    s = val.strip().lower()
    for h in ALLOWED_HUBS:
        if h in s:
            return True
    return False

def robust_read_csv(csv_file):
    try:
        df = pd.read_csv(csv_file)
    except Exception:
        csv_file.seek(0)
        df = pd.read_csv(csv_file, encoding="latin1")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def make_imgsrc_map(product_df):
    img_map = {}
    for _, row in product_df.iterrows():
        pid = str(row['MB_id']).strip()
        img_src = str(row['image_src']).strip()
        if pid and img_src and pid.lower() != 'nan' and img_src.lower() != 'nan':
            img_map[pid] = img_src
    return img_map

def mk_mb_img_link(img_src):
    if not img_src: return ""
    return f"https://file.milkbasket.com/products/{img_src}"

def mk_amzid_link(img_src):
    if not img_src: return ""
    img_src_png = re.sub(r'\.\w+$', '.png', img_src)
    return f"https://design-figma.s3.ap-south-1.amazonaws.com/{img_src_png}"

def match_col(cols, name):
    name_clean = re.sub(r'\s+', '', name).lower()
    for col in cols:
        col_clean = re.sub(r'\s+', '', col).lower()
        if col_clean == name_clean:
            return col
    return None

def process_kvi(df, img_map):
    cols = df.columns
    pname_col = match_col(cols, "Product Name")
    mbid1_col = match_col(cols, "MB ID 1")
    subcat_col = match_col(cols, "Focused Sub Cat")
    out = []
    for idx, row in df.iterrows():
        hub = str(row.iloc[0]).strip()
        # if not is_hub(hub): continue
        mbid1 = clean_mbid(row.get(mbid1_col, "")) if mbid1_col else ""
        if not mbid1 or mbid1.lower() == "nan": continue
        pname = row.get(pname_col, "") if pname_col else ""
        subcat = row.get(subcat_col, "") if subcat_col else ""
        img_src1 = img_map.get(mbid1, "") if img_map else ""
        out.append({
            "Hub": hub,
            "Product Name": pname,
            "MB ID 1": mbid1,
            "Focused Sub Cat": subcat,
            "Img1": mk_mb_img_link(img_src1),
            "Img2": "",
            "AmzId1": mk_amzid_link(img_src1),
            "AmzId2": ""
        })
    return out

def process_range_dualmrp(df, img_map, add_copy):
    cols = df.columns
    pname_col = match_col(cols, "Product Name")
    mbid1_col = match_col(cols, "MB ID 1")
    mbid2_col = match_col(cols, "MB ID 2")
    subcat_col = match_col(cols, "Focused Sub Cat")
    callout_col = match_col(cols, "Banner Call-Out") 
    out = []
    for idx, row in df.iterrows():
        hub = str(row.iloc[0]).strip()
        # if not is_hub(hub): continue
        mbid1 = clean_mbid(row.get(mbid1_col, "")) if mbid1_col else ""
        mbid2 = clean_mbid(row.get(mbid2_col, "")) if mbid2_col else ""
        if (not mbid1 or mbid1.lower() == "nan") and (not mbid2 or mbid2.lower() == "nan"):
            continue
        pname = row.get(pname_col, "") if pname_col else ""
        subcat = row.get(subcat_col, "") if subcat_col else ""
        callout = row.get(callout_col, "") if callout_col else "" 
        img_src1 = img_map.get(mbid1, "") if img_map else ""
        img_src2 = img_map.get(mbid2, "") if img_map else ""
        outrow = {
            "Hub": hub,
            "Product Name": pname,
            "MB ID 1": mbid1,
            "MB ID 2": mbid2,
            "Focused Sub Cat": subcat,
            "Banner Call-Out": callout, 
            "Copy": "",  # Blank for user to fill
            "Img1": mk_mb_img_link(img_src1),
            "Img2": mk_mb_img_link(img_src2),
            "AmzId1": mk_amzid_link(img_src1),
            "AmzId2": mk_amzid_link(img_src2)
        }
        out.append(outrow)
    return out

def clean_sheet_name(name):
    return re.sub(r'[\[\]\*:/\\?]', '', str(name)).strip()[:31]

def create_all_mbids_tab(tab_dict, img_map):
    mbid_set = set()
    for tab in tab_dict.values():
        for row in tab:
            for mb_col in ["MB ID 1", "MB ID 2"]:
                mbid = row.get(mb_col, "")
                if mbid and mbid.lower() != "nan":
                    mbid_set.add(str(mbid))  # <--- always cast to str here!
    all_mbids = sorted(str(x) for x in mbid_set)
    rows = []
    for mbid in all_mbids:
        img_src = img_map.get(mbid, "")
        img_link = mk_mb_img_link(img_src) if img_src else ""
        rows.append({"MB ID": mbid, "Image Link": img_link})
    return rows

def excel_export(tab_dict, all_mbids_tab):
    output = BytesIO()
    wb = Workbook()
    for tname, rows in tab_dict.items():
        ws = wb.create_sheet(title=clean_sheet_name(tname))
        df = pd.DataFrame(rows)
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        bold_font = Font(bold=True)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font
    # All_MbIDs tab with red highlights for missing links
    ws2 = wb.create_sheet("All_MbIDs")
    df_mbids = pd.DataFrame(all_mbids_tab)
    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    for r_idx, row in enumerate(dataframe_to_rows(df_mbids, index=False, header=True), 1):
        ws2.append(row)
        if r_idx > 1:
            img_link = row[1] if len(row) > 1 else ""
            if not img_link or img_link.strip() == "":
                cell = ws2.cell(row=r_idx, column=2)
                cell.fill = red_fill
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output)
    output.seek(0)
    return output

def collect_all_images(tab_dict):
    images = {}
    for tab in tab_dict:
        for row in tab_dict[tab]:
            for col, pid, linkcol in [("Img1", "MB ID 1", "Img1"), ("Img2", "MB ID 2", "Img2")]:
                link = row.get(linkcol, "")
                pid_val = row.get(pid, "")
                if link and pid_val:
                    images[pid_val] = link
    return images

def has_transparency(img_bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            alpha = img.getchannel("A") if "A" in img.getbands() else None
            if alpha and alpha.getextrema()[0] < 255:
                return True
        return False
    except Exception:
        return False

# --- UI ---
uploaded_excel = st.file_uploader("Upload Banners Excel", type=["xlsx"])
product_csv = st.file_uploader("Upload product CSV (with 'MB_id' and 'image_src') (optional)", type=["csv"])

img_map = {}
if product_csv:
    product_df = robust_read_csv(product_csv)
    if not {"MB_id", "image_src"}.issubset(product_df.columns):
        st.error("CSV must have columns 'MB_id' and 'image_src'.")
        st.stop()
    img_map = make_imgsrc_map(product_df)

tab_dict = {}
if uploaded_excel:
    xls = pd.ExcelFile(uploaded_excel)
    for tab in xls.sheet_names:
        tab_low = tab.lower().strip()
        if tab_low == "kvi":
            df = pd.read_excel(xls, sheet_name=tab)
            tab_dict["KVI"] = process_kvi(df, img_map)
        elif tab_low == "range":
            df = pd.read_excel(xls, sheet_name=tab)
            tab_dict["Range"] = process_range_dualmrp(df, img_map, add_copy=True)
        elif tab_low == "dual mrp":
            df = pd.read_excel(xls, sheet_name=tab)
            tab_dict["Dual MRP"] = process_range_dualmrp(df, img_map, add_copy=True)

# --- Output and Images ---
if tab_dict:
    all_mbids_tab = create_all_mbids_tab(tab_dict, img_map)
with st.expander("Preview All_MBIDs tab"):
    st.dataframe(pd.DataFrame(all_mbids_tab))
    st.success("✅ Processed! Preview tabs below, then download the full Excel output.")
    tab_names = list(tab_dict.keys()) + ["All_MbIDs"]
    selected_tab = st.selectbox("Preview a tab:", tab_names)
    if selected_tab == "All_MbIDs":
        st.dataframe(pd.DataFrame(all_mbids_tab))
    else:
        st.dataframe(pd.DataFrame(tab_dict[selected_tab]))
    output = excel_export(tab_dict, all_mbids_tab)
    st.download_button(
        "📥 Download Multi-Tab Banners Excel",
        output,
        file_name="Banners_MultiTab_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
with st.expander("Preview All_PIDs Tab"):
    st.dataframe(pd.DataFrame(all_pids_tab))  
    
    # Image Download Section
    all_images = collect_all_images(tab_dict)
    st.markdown("## Download all banner product images")
    if st.button("Download Images.zip"):
        progress = st.progress(0)
        total = len(all_images)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for idx, (pid, link) in enumerate(all_images.items()):
                try:
                    r = requests.get(link, timeout=10)
                    if r.status_code == 200:
                        img = Image.open(io.BytesIO(r.content)).convert("RGBA")
                        img_byte_arr = io.BytesIO()
                        img.save(img_byte_arr, format='PNG')
                        zipf.writestr(f"{pid}.png", img_byte_arr.getvalue())
                except Exception:
                    continue
                progress.progress((idx + 1) / total)
        zip_buffer.seek(0)
        st.success("Images.zip ready for download!")
        st.download_button(
            label="Download Images.zip",
            data=zip_buffer,
            file_name="Banner_Images.zip",
            mime="application/zip"
        )

    st.markdown("## Download images with background removed (rembg)")
if st.button("Download rembg Images.zip"):
        progress = st.progress(0)
        total = len(all_images)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for idx, (pid, link) in enumerate(all_images.items()):
                try:
                    r = requests.get(link, timeout=10)
                    if r.status_code == 200:
                        img = Image.open(io.BytesIO(r.content)).convert("RGBA")
                        if has_transparency(r.content):
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            zipf.writestr(f"{pid}.png", img_byte_arr.getvalue())
                        else:
                            img = img.resize((650, 650))
                            output_image = remove(img)
                            img_byte_arr = io.BytesIO()
                            output_image.save(img_byte_arr, format='PNG')
                            zipf.writestr(f"{pid}.png", img_byte_arr.getvalue())
                except Exception:
                    continue
                progress.progress((idx + 1) / total)
        zip_buffer.seek(0)
        st.success("rembg Images.zip ready for download!")
        st.download_button(
            label="Download rembg Images.zip",
            data=zip_buffer,
            file_name="Banner_Images_rembg.zip",
            mime="application/zip"
        )
else:
    st.info("Please upload your Banners Excel (and optionally the product CSV).")
